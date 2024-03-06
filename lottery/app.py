from flask import Flask, render_template, request, redirect, url_for
from openpyxl import Workbook, load_workbook
import random
import os

app = Flask(__name__)

participants_file = "participants.xlsx"


def reset_participants():
    if os.path.exists(participants_file):
        os.remove(participants_file)
    wb = Workbook()
    ws = wb.active
    ws.append(["Name"])
    wb.save(participants_file)


reset_participants()


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/add_name', methods=['POST'])
def add_name():
    name = request.form['name']
    wb = load_workbook(participants_file)
    ws = wb.active
    ws.append([name])
    wb.save(participants_file)
    return redirect(url_for('index'))


@app.route('/draw_winner', methods=['GET', 'POST'])
def draw_winner():
    if request.method == 'POST':
        num_winners = int(request.form['num_winners'])
        num_participants = load_workbook(
            participants_file).active.max_row - 1
        if num_winners > num_participants:
            return "Error: Number of winners exceeds number of participants."
        elif num_winners <= 0:
            return "Error: Number of winners must be positive."

        participants = []
        wb = load_workbook(participants_file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1, values_only=True):
            participants.append(row[0])

        winners = random.sample(participants, num_winners)
        return render_template('winners.html', winners=winners)
    else:
        return render_template('draw.html')


if __name__ == '__main__':
    app.run(debug=True)
